import os
import time
import json
import base64
import mimetypes
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, Optional

from dotenv import load_dotenv
import anthropic
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from natsort import natsorted

from config_ereceipt import (
    HEADERS, DATE_FMT, CURRENCY_FMT,
    FIELD_MAPPING, SCHEMA, SYSTEM_PROMPT, USER_PROMPT, SHEET_CONFIGS
)

load_dotenv()


def file_to_base64(file_path: str) -> str:
    """Read a file and return its base64-encoded content."""
    with open(file_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


def safe_decimal(x: Any) -> Optional[Decimal]:
    if x is None:
        return None
    try:
        return Decimal(str(x))
    except (InvalidOperation, ValueError):
        return None


def ensure_sheet_and_headers(wb_path: str, headers: list[str], sheet_name: str) -> tuple[Any, Worksheet]:
    """
    Creates/opens the workbook and ensures the target sheet + headers exist.
    """
    if os.path.exists(wb_path):
        wb = load_workbook(wb_path)
    else:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
        ws.title = sheet_name

    # Write headers if missing / mismatched
    existing = [ws.cell(row=1, column=i + 1).value for i in range(len(headers))]
    if existing != headers:
        for i, h in enumerate(headers, start=1):
            ws.cell(row=1, column=i).value = h

    return wb, ws


def append_row(
    ws: Worksheet,
    extracted: Dict[str, Any],
    field_mapping: list[tuple],
    date_fmt: str,
    currency_fmt: str,
) -> None:
    """
    Appends a row based on field mapping configuration.
    """
    next_row = ws.max_row + 1

    for header, json_key, col_idx, format_type in field_mapping:
        cell = ws.cell(row=next_row, column=col_idx)

        if json_key is None:
            # Blank cell (e.g., Entertainment)
            cell.value = None
        elif format_type == "date":
            # Parse date
            date_str = extracted.get(json_key)
            if date_str:
                try:
                    cell.value = datetime.strptime(date_str, "%Y-%m-%d").date()
                    cell.number_format = date_fmt
                except ValueError:
                    cell.value = None
            else:
                cell.value = None
        elif format_type == "currency":
            # Currency with formatting
            val = extracted.get(json_key)
            decimal_val = safe_decimal(val)
            cell.value = float(decimal_val) if decimal_val is not None else None
            cell.number_format = currency_fmt
        else:
            # Text
            val = extracted.get(json_key)
            cell.value = (val or "").strip() or None if isinstance(val, str) else val


def extract_receipt_fields(
    client: anthropic.Anthropic,
    model: str,
    schema: Dict[str, Any],
    system_prompt: str,
    user_prompt: str,
    file_path: str,
) -> Dict[str, Any]:
    """
    Calls the Anthropic Messages API with:
      - image or PDF input (base64)
      - Structured Outputs (JSON schema) to force consistent fields

    Claude handles PDFs natively — no need for separate text extraction
    or image conversion.
    """
    b64_data = file_to_base64(file_path)
    ext = os.path.splitext(file_path.lower())[1]

    # Build the content block based on file type
    if ext == ".pdf":
        file_block = {
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": "application/pdf",
                "data": b64_data,
            },
        }
    else:
        mime, _ = mimetypes.guess_type(file_path)
        if not mime:
            mime = "image/jpeg"
        file_block = {
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": mime,
                "data": b64_data,
            },
        }

    resp = client.messages.create(
        model=model,
        max_tokens=1024,
        system=system_prompt,
        messages=[
            {
                "role": "user",
                "content": [
                    file_block,
                    {"type": "text", "text": user_prompt},
                ],
            },
        ],
        output_config={
            "format": {
                "type": "json_schema",
                "schema": schema,
            }
        },
    )

    raw = (resp.content[0].text or "").strip()
    if not raw:
        return {key: None for key in schema.get("required", [])}

    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return {key: None for key in schema.get("required", [])}


def get_files(folder: str = "receipt_images", limit: Optional[int] = None) -> list[str]:
    """Get image and PDF files from the specified folder."""
    if not os.path.exists(folder):
        raise FileNotFoundError(f"Folder '{folder}' not found.")

    supported_exts = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".pdf"}
    files: list[str] = []
    for filename in natsorted(os.listdir(folder), reverse=True):
        if os.path.splitext(filename.lower())[1] in supported_exts:
            files.append(os.path.join(folder, filename))
            if limit and len(files) >= limit:
                break

    if not files:
        raise FileNotFoundError(f"No image/PDF files found in '{folder}'.")

    return files


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Extract receipt fields from images/PDFs and append to an .xlsx sheet.")
    parser.add_argument("-n", "--num", type=int, default=0, help="Number of files to process per folder (default: 0 for all).")
    parser.add_argument("--xlsx", default="receipts.xlsx", help="Output workbook path (default: receipts.xlsx).")
    parser.add_argument("--model", default=os.getenv("RECEIPT_MODEL") or "claude-sonnet-4-5-20250514")
    parser.add_argument("--folder", help="Process a single folder (provide folder path). If omitted, processes all folders from config.")
    parser.add_argument("--sheet", help="Sheet name (required if --folder is specified).")
    args = parser.parse_args()

    client = anthropic.Anthropic()
    limit = None if args.num == 0 else args.num

    # Single folder mode or batch mode
    if args.folder:
        if not args.sheet:
            print("Error: --sheet is required when using --folder")
            return
        configs = [{"folder": args.folder, "sheet_name": args.sheet}]
    else:
        configs = SHEET_CONFIGS

    # Open/create workbook once
    is_new_wb = False
    if os.path.exists(args.xlsx):
        wb = load_workbook(args.xlsx)
    else:
        wb = Workbook()
        is_new_wb = True

    total_processed = 0
    sheets_created = 0

    for config in configs:
        folder = config["folder"]
        sheet_name = config["sheet_name"]

        try:
            files = get_files(folder=folder, limit=limit)
        except FileNotFoundError as e:
            print(f"Skipping {folder}: {e}")
            continue

        print(f"\nProcessing {len(files)} file(s) from '{folder}' -> sheet '{sheet_name}'...")

        # Ensure sheet and headers
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            sheets_created += 1
            # If this is a new workbook, remove the default 'Sheet' now that we have a real one
            if is_new_wb and "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
                is_new_wb = False

        # Write headers
        existing = [ws.cell(row=1, column=i + 1).value for i in range(len(HEADERS))]
        if existing != HEADERS:
            for i, h in enumerate(HEADERS, start=1):
                ws.cell(row=1, column=i).value = h

        for i, file_path in enumerate(files, 1):
            filename = os.path.basename(file_path)
            print(f"  [{i}/{len(files)}] Processing: {filename}")

            user_prompt = USER_PROMPT

            start_time = time.time()
            extracted = extract_receipt_fields(
                client,
                args.model,
                SCHEMA,
                SYSTEM_PROMPT,
                user_prompt,
                file_path,
            )
            duration = time.time() - start_time
            print(f"    Done in {duration:.2f}s")

            append_row(ws, extracted, FIELD_MAPPING, DATE_FMT, CURRENCY_FMT)
            total_processed += 1

    # Only save if we actually have sheets to save (at least one must be visible)
    try:
        if len(wb.sheetnames) > 0:
            wb.save(args.xlsx)
            print(f"\n✓ Total: {total_processed} receipt(s) processed across {len(configs)} sheet(s) in {args.xlsx}")
        else:
            print("\n⚠ No sheets were created because no files were found. Workbook not saved.")
    except (IndexError, Exception) as e:
        print(f"\n⚠ Error: Could not save workbook: {e}")


if __name__ == "__main__":
    main()
